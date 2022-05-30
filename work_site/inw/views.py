from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.urls import reverse_lazy
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from datetime import datetime
from .models import InwModel, UploadModel
from .forms import CreateDataForm, SurplusLackInputForm, UploadModelFormSelect, EditForm
from django.views.generic import View, UpdateView, CreateView
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin


# Create your views here.


# REDIRECT PREVIOUS PAGE METHOD FOR GENERIC CLASS VIEW

class RedirectToPreviousMixin:
    default_redirect = '/'

    def get(self, request, *args, **kwargs):
        request.session['previous_page'] = request.META.get('HTTP_REFERER', self.default_redirect)
        return super().get(request, *args, **kwargs)

    def get_success_url(self):
        return self.request.session['previous_page']


@login_required(login_url="/login")
def download_data_as_excel(request, pk):
    if request.method == 'GET':
        model_queryset = InwModel.objects.filter(upload=pk)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f"attachment; filename={datetime.now().strftime('%Y-%m-%d')}-akcesoria.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Akcesoria'
        columns = [
            'ID',
            'Nazwa',
            'EAN',
            'Ilość',
        ]
        row_num = 1
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            wsc = worksheet[cell.coordinate]
            cell.value = column_title
            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            wsc.font = Font(bold=True, size="13")
            wsc.alignment = Alignment(horizontal='center', vertical='center')
            worksheet.row_dimensions[1].height = 35
            if column_title == 'ID' or column_title == 'Ilość':
                column_dimensions.width = 8
            elif column_title == 'EAN':
                column_dimensions.width = 30
            else:
                column_dimensions.width = 40
        for model in model_queryset:
            row_num += 1
            row = [
                row_num - 1,
                model.name,
                model.EAN,
                model.quantity,
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                wsc = worksheet[cell.coordinate]
                if cell.column_letter == 'A':
                    wsc.alignment = Alignment(horizontal='left')
                elif cell.column_letter == 'C':
                    wsc.alignment = Alignment(horizontal='center')
                elif cell.column_letter == 'D':
                    if cell_value > 0:
                        wsc.font = Font(color="0000FF00")
                    elif cell_value < 0:
                        wsc.font = Font(color="00FF0000")
                cell.value = cell_value
        worksheet.auto_filter.ref = worksheet.dimensions
        workbook.save(response)
        return response


class ConfirmDeleteList(LoginRequiredMixin, View):
    login_url = '/login/'

    def get(self, request, *args, **kwargs):
        request.session['previous_page'] = request.META.get('HTTP_REFERER', '/')
        id_list = request.GET.getlist('delete')
        product = InwModel
        if len(id_list) == 0:
            messages.add_message(request, messages.ERROR, 'Musisz wybrać pozycję żeby usunąć!')
            return redirect(request.session['previous_page'])
        else:
            objects = product.objects.filter(id__in=id_list)
            return render(request, 'inw/inwmodel_delete_list.html', {'objects': objects})

    def post(self, request, *args, **kwargs):
        id_list = request.GET.getlist('delete')
        product = InwModel
        container = UploadModel
        product_id = product.objects.get(pk=id_list[0])
        field_name_val = getattr(product_id, 'upload')
        product.objects.filter(id__in=id_list).delete()
        field_len = len(field_name_val.set.all())
        if field_len == 0:
            container.objects.get(id=field_name_val.id).delete()
            messages.add_message(request, messages.SUCCESS, f'Usunięto wszystkie obiekty w {field_name_val}!')
            return redirect(reverse('myapp:table'))
        messages.add_message(request, messages.SUCCESS, f'Usunięto {len(id_list)} elementów!')
        return redirect(request.session['previous_page'])


class UploadData(LoginRequiredMixin, View):
    login_url = '/login/'

    def get(self, request):
        return render(request, 'inw/upload_form_page.html')

    def post(self, request, *args, **kwargs):
        from .upload_scripts.upload_scripts import excel_inf_to_list, excel_sap_to_dict, process_excel_files, \
            handle_uploaded
        from .upload_scripts.user_query_check_NaN import check_NaN
        try:
            sap, inw = handle_uploaded(request.FILES)
            user = request.user
            df_sap = pd.read_excel(sap)
            df_inw = pd.read_excel(inw, header=None)
            inw_list = excel_inf_to_list(df_inw, df_sap)
            sap_dict = excel_sap_to_dict(df_sap)
            new_data = process_excel_files(inw_list, sap_dict)
            user_query = UploadModel.objects.filter(user=user)
            name = check_NaN(user_query, new_data)
            upload_file = UploadModel(name=name, user=user)
            upload_file.save()
            sql_data = {
                'name': '',
                'EAN': '',
                'quantity': '',
                'upload': upload_file
            }
            index = -1
            for value in new_data['Zapas ogółem']:
                index += 1
                if value != 0:
                    sql_data['quantity'] = new_data['Zapas ogółem'][index]
                    sql_data['EAN'] = new_data['EAN'][index]
                    sql_data['name'] = new_data['Krótki tekst materiału'][index]
                    model = InwModel(**sql_data)
                    model.save()
            id = getattr(upload_file, 'id')
            link = reverse('myapp:table')
            response = {'url': link}
            messages.add_message(request, messages.SUCCESS, 'Plik został przesłany!')
            return JsonResponse(response)
        except:
            messages.add_message(request, messages.ERROR, 'Coś poszło nie tak')
            link = reverse('myapp:upload')
            response = {'url': link}
            return JsonResponse(response)


class TableData(LoginRequiredMixin, View):
    login_url = '/login/'

    def get(self, request, **kwargs):
        checkboxesform = SurplusLackInputForm()
        selectform = UploadModelFormSelect(user=request.user)
        ProductModel = InwModel
        productContainerModel = UploadModel
        context = {
            'checkboxesform': checkboxesform,
            'selectform': selectform,
        }
        intaial_data = {

        }
        try:
            if checkboxesform.is_valid and selectform.is_valid:
                productcontainer = productContainerModel.objects.get(id=request.GET['upload'])
                products = ProductModel.objects.filter(upload=productcontainer)
                intaial_data['upload'] = productcontainer
                context['selectform'] = UploadModelFormSelect(initial=intaial_data, user=request.user)
                if 'surplus_check' in request.GET and 'lack_check' not in request.GET:
                    checkboxesform.fields['lack_check'].initial = False
                    products = products.filter(quantity__gt=0)
                    context["products"] = products
                elif 'lack_check' in request.GET and 'surplus_check' not in request.GET:
                    checkboxesform.fields['surplus_check'].initial = False
                    products = products.filter(quantity__lt=0)
                    context["products"] = products
                else:
                    context["products"] = products
        finally:
            return render(request, 'inw/table_form.html', context=context)


class InwModelCreateView(RedirectToPreviousMixin, LoginRequiredMixin, SuccessMessageMixin, CreateView):
    form_class = CreateDataForm
    model = InwModel
    template_name_suffix = '_create_form'
    success_url = reverse_lazy('myapp:table')
    success_message = "Obiekt %(name)s został utworzony!"
    login_url = '/login/'

    def get_form(self, *args, **kwargs):
        form = super(InwModelCreateView, self).get_form(*args, **kwargs)
        form.fields['upload'].queryset = UploadModel.objects.filter(user=self.request.user)
        return form


class InwModelUpdateView(RedirectToPreviousMixin, LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    form_class = EditForm
    model = InwModel
    template_name_suffix = '_update_form'
    success_url = reverse_lazy('myapp:table')
    success_message = f"Nowa wartość %(quantity)s została dodana!"
    login_url = '/login/'
